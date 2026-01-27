import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime


LENGTHS_ML: List[str] = [
    "1.0", "3.5", "4.0", "4.5", "5.0", "5.5", "6.0", "6.5", "7.0", "7.5",
    "8.0", "8.5", "9.0", "9.5", "10.0", "10.5", "11.0", "11.5", "12.0", "12.5", "13.0",
]


def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "").replace(" ", "")
    try:
        return float(s)
    except Exception:
        return None


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.strip()


def export_excel(json_path: str, xlsx_path: str) -> None:
    """
    Export the optimized JSON cost matrix into a human-friendly Excel workbook.

    Canonical editable sheet: PRODUCTS
    View sheets: INDEX, COSTOS, PRECIOS
    Metadata sheet: METADATA
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    json_path_p = Path(json_path)
    data = json.loads(json_path_p.read_text(encoding="utf-8"))

    products: List[Dict[str, Any]] = data.get("productos", {}).get("todos", [])

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical="center")

    def write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
        ws.append(headers)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Basic column sizing (heuristic)
        for col_idx, h in enumerate(headers, start=1):
            width = min(max(len(h) + 2, 12), 40)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # PRODUCTS (canonical)
    ws_products = wb.create_sheet("PRODUCTS")
    headers_products = [
        "proveedor",
        "codigo",
        "nombre",
        "categoria",
        "espesor_mm",
        "estado",
        "shopify_status",
        "notas",
        "costo_base_usd_iva",
        "costo_con_aumento_usd_iva",
        "costo_proximo_aumento_usd_iva",
        "margen_porcentaje",
        "ganancia_usd",
        "precio_empresa_venta_iva_usd",
        "precio_particular_consumidor_iva_inc_usd",
        "precio_web_venta_iva_usd",
        "precio_web_venta_iva_inc_usd",
        "precio_ml_base_usd",
    ] + [f"ml_{l}" for l in LENGTHS_ML]

    rows_products: List[List[Any]] = []
    for p in products:
        meta = p.get("metadata", {}) or {}
        costos = (p.get("costos", {}) or {}).get("fabrica_directo", {}) or {}
        margen = p.get("margen", {}) or {}
        precios = p.get("precios", {}) or {}
        empresa = (precios.get("empresa", {}) or {})
        particular = (precios.get("particular", {}) or {})
        web_stock = (precios.get("web_stock", {}) or {})
        ml = p.get("precio_metro_lineal", {}) or {}
        ml_by_len = (ml.get("precios_por_largo", {}) or {})

        row = [
            _safe_str(meta.get("proveedor")),
            _safe_str(p.get("codigo")),
            _safe_str(p.get("nombre")),
            _safe_str(p.get("categoria")),
            _safe_str(p.get("espesor_mm")),
            _safe_str(p.get("estado")),
            _safe_str(meta.get("shopify_status")),
            _safe_str(meta.get("notas")),
            costos.get("costo_base_usd_iva"),
            costos.get("costo_con_aumento_usd_iva"),
            costos.get("costo_proximo_aumento_usd_iva"),
            _safe_str(margen.get("porcentaje")),
            margen.get("ganancia_usd"),
            empresa.get("venta_iva_usd"),
            particular.get("consumidor_iva_inc_usd"),
            web_stock.get("web_venta_iva_usd"),
            web_stock.get("web_venta_iva_inc_usd"),
            ml.get("precio_base_usd"),
        ]
        for l in LENGTHS_ML:
            row.append(ml_by_len.get(l))
        rows_products.append(row)

    write_table(ws_products, headers_products, rows_products)

    # INDEX (view)
    ws_index = wb.create_sheet("INDEX")
    headers_index = [
        "proveedor",
        "codigo",
        "nombre",
        "categoria",
        "espesor_mm",
        "estado",
        "precio_empresa_venta_iva_usd",
        "precio_particular_consumidor_iva_inc_usd",
        "precio_web_venta_iva_usd",
        "precio_web_venta_iva_inc_usd",
    ]
    rows_index = []
    for r in rows_products:
        # based on PRODUCTS header ordering
        proveedor, codigo, nombre, categoria, espesor, estado = r[0], r[1], r[2], r[3], r[4], r[5]
        p_emp = r[13]
        p_part = r[14]
        p_web = r[15]
        p_web_inc = r[16]
        rows_index.append([proveedor, codigo, nombre, categoria, espesor, estado, p_emp, p_part, p_web, p_web_inc])
    write_table(ws_index, headers_index, rows_index)

    # COSTOS (view)
    ws_costos = wb.create_sheet("COSTOS")
    headers_costos = [
        "proveedor",
        "codigo",
        "nombre",
        "categoria",
        "estado",
        "costo_base_usd_iva",
        "costo_con_aumento_usd_iva",
        "costo_proximo_aumento_usd_iva",
        "margen_porcentaje",
        "ganancia_usd",
    ]
    rows_costos = []
    for r in rows_products:
        rows_costos.append([r[0], r[1], r[2], r[3], r[5], r[8], r[9], r[10], r[11], r[12]])
    write_table(ws_costos, headers_costos, rows_costos)

    # PRECIOS (view)
    ws_precios = wb.create_sheet("PRECIOS")
    headers_precios = [
        "proveedor",
        "codigo",
        "nombre",
        "categoria",
        "estado",
        "precio_empresa_venta_iva_usd",
        "precio_particular_consumidor_iva_inc_usd",
        "precio_web_venta_iva_usd",
        "precio_web_venta_iva_inc_usd",
        "precio_ml_base_usd",
    ]
    rows_precios = []
    for r in rows_products:
        rows_precios.append([r[0], r[1], r[2], r[3], r[5], r[13], r[14], r[15], r[16], r[17]])
    write_table(ws_precios, headers_precios, rows_precios)

    # METADATA
    ws_meta = wb.create_sheet("METADATA")
    ws_meta.append(["key", "value"])
    ws_meta["A1"].font = header_font
    ws_meta["B1"].font = header_font
    ws_meta.freeze_panes = "A2"
    meta = data.get("meta", {}) or {}
    ws_meta.append(["generated_at", datetime.now().isoformat()])
    for k, v in meta.items():
        if isinstance(v, (dict, list)):
            ws_meta.append([k, json.dumps(v, ensure_ascii=False)])
        else:
            ws_meta.append([k, v])

    # Save
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def import_excel(xlsx_path: str, output_json_path: str, base_json_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Import the Excel workbook back into the optimized JSON format.

    Reads canonical data from the PRODUCTS sheet.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    if "PRODUCTS" not in wb.sheetnames:
        raise ValueError("Missing required sheet: PRODUCTS")

    ws = wb["PRODUCTS"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty PRODUCTS sheet")

    headers = [_safe_str(h) for h in rows[0]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    def get(row, col, default=None):
        i = header_idx.get(col)
        if i is None or i >= len(row):
            return default
        return row[i]

    products: List[Dict[str, Any]] = []
    for r in rows[1:]:
        codigo = _safe_str(get(r, "codigo"))
        nombre = _safe_str(get(r, "nombre"))
        if not codigo or not nombre:
            continue

        proveedor = _safe_str(get(r, "proveedor"))
        categoria = _safe_str(get(r, "categoria"))
        espesor_mm = _safe_str(get(r, "espesor_mm")) or None
        estado = _safe_str(get(r, "estado")) or "ACT."
        shopify_status = _safe_str(get(r, "shopify_status"))
        notas = _safe_str(get(r, "notas"))

        costos = {
            "fabrica_directo": {
                "costo_base_usd_iva": _num(get(r, "costo_base_usd_iva")),
                "costo_con_aumento_usd_iva": _num(get(r, "costo_con_aumento_usd_iva")),
                "costo_proximo_aumento_usd_iva": _num(get(r, "costo_proximo_aumento_usd_iva")),
            }
        }
        margen = {
            "porcentaje": _safe_str(get(r, "margen_porcentaje")),
            "ganancia_usd": _num(get(r, "ganancia_usd")),
        }
        precios = {
            "empresa": {
                "venta_iva_usd": _num(get(r, "precio_empresa_venta_iva_usd")),
                "nota": "Empresas descuentan IVA, usar precio + IVA",
            },
            "particular": {
                "consumidor_iva_inc_usd": _num(get(r, "precio_particular_consumidor_iva_inc_usd")),
                "nota": "Particulares no descuentan IVA, usar precio IVA incluido",
            },
            "web_stock": {
                "web_venta_iva_usd": _num(get(r, "precio_web_venta_iva_usd")),
                "web_venta_iva_inc_usd": _num(get(r, "precio_web_venta_iva_inc_usd")),
            },
        }

        ml_by_len: Dict[str, Any] = {}
        for l in LENGTHS_ML:
            v = _num(get(r, f"ml_{l}"))
            if v is not None:
                ml_by_len[l] = v

        precio_metro_lineal = {
            "precio_base_usd": _num(get(r, "precio_ml_base_usd")),
            "precios_por_largo": ml_by_len,
        }

        product = {
            "codigo": codigo,
            "nombre": nombre,
            "categoria": categoria or "otros",
            "espesor_mm": espesor_mm,
            "estado": estado,
            "costos": costos,
            "margen": margen,
            "precios": precios,
            "precio_metro_lineal": precio_metro_lineal,
            "metadata": {
                "proveedor": proveedor,
                "shopify_status": shopify_status,
                "notas": notas,
                "fecha_actualizacion": datetime.now().isoformat(),
                "source": "excel_import",
            },
        }
        products.append(product)

    # Base structure
    if base_json_path:
        base = json.loads(Path(base_json_path).read_text(encoding="utf-8"))
    else:
        base = {
            "meta": {
                "nombre": "Cost Matrix (Imported from Excel)",
                "version": "2.0.0",
                "fecha_creacion": datetime.now().isoformat(),
                "optimizado_para": "GPT OpenAI Actions - KB Indexing",
            },
            "reglas_precios": {},
        }

    # Recompute indexes using redesign tool logic if available
    from .redesign_tool import CostMatrixRedesigner
    redesigner = CostMatrixRedesigner()
    structure = redesigner.create_optimized_structure(products)

    # Preserve base meta fields where sensible
    structure["meta"].update({k: v for k, v in base.get("meta", {}).items() if k not in {"fecha_creacion", "total_productos", "total_categorias", "estadisticas"}})
    if base.get("reglas_precios"):
        structure["reglas_precios"] = base["reglas_precios"]

    Path(output_json_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_json_path).write_text(json.dumps(structure, indent=2, ensure_ascii=False), encoding="utf-8")
    return structure


def main() -> None:
    import sys

    if len(sys.argv) < 2:
        print(
            "Usage:\n"
            "  python -m panelin_improvements.cost_matrix_tools.excel_manager export <input_json> <output_xlsx>\n"
            "  python -m panelin_improvements.cost_matrix_tools.excel_manager import <input_xlsx> <output_json> [base_json]\n"
        )
        raise SystemExit(1)

    cmd = sys.argv[1].lower()
    if cmd == "export":
        if len(sys.argv) < 4:
            raise SystemExit("export requires <input_json> <output_xlsx>")
        export_excel(sys.argv[2], sys.argv[3])
        print("Done!")
        return

    if cmd == "import":
        if len(sys.argv) < 4:
            raise SystemExit("import requires <input_xlsx> <output_json> [base_json]")
        base = sys.argv[4] if len(sys.argv) >= 5 else None
        import_excel(sys.argv[2], sys.argv[3], base_json_path=base)
        print("Done!")
        return

    raise SystemExit(f"Unknown command: {cmd}")


if __name__ == "__main__":
    main()

